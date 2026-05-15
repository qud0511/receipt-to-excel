"""Phase 6.8 — Templates API.

ADR-006/011 휴리스틱 + ADR-010 추천 3 (Phase 6.8 셀 값 + 매핑 chips PATCH 만,
style/병합/줌은 Phase 8+).

CLAUDE.md 강제: 모든 변경 라우터에 Depends(get_current_user) + IDOR 차단.
"""

from __future__ import annotations

import io
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Annotated, Any

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Request,
    UploadFile,
    status,
)
from fastapi.responses import FileResponse

# 부채(P8.12 발견): CLAUDE.md "외부 의존성은 services/* 뒤에만 — api 에서 openpyxl
# 직접 import 금지" 위반. import-linter openpyxl-forbidden contract 는 본 라인을
# services 인터페이스 뒤로 옮기는 별도 sub-phase 리팩터 후 활성화 예정.
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.security import UploadGuard, UploadValidationError
from app.db.repositories import template_repo, user_repo
from app.schemas.auth import UserInfo
from app.schemas.template import (
    AnalyzedTemplateResponse,
    CellsPatchRequest,
    GridCell,
    GridResponse,
    GridSheetView,
    MappingPatchRequest,
    MetaPatchRequest,
    TemplateCreatedResponse,
    TemplateSummary,
    sheet_config_to_view,
)
from app.services.templates.analyzer import TemplateAnalysisError, analyze_workbook

router = APIRouter(prefix="/templates", tags=["templates"])


async def _get_db(request: Request) -> AsyncIterator[AsyncSession]:
    sessionmaker = request.app.state.db_sessionmaker
    async with sessionmaker() as session:
        yield session


# ── 1) GET /templates — list ──────────────────────────────────────────────────
@router.get("", response_model=list[TemplateSummary])
async def list_templates(
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> list[TemplateSummary]:
    """사용자별 list. UI Templates sidebar 의 카드 list."""
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    templates = await template_repo.list_for_user(db, user_id=db_user.id)
    return [
        TemplateSummary(
            id=t.id,
            name=t.name,
            is_default=t.is_default,
            mapping_status=t.mapping_status,
            created_at=t.created_at,
            updated_at=t.updated_at,
        )
        for t in templates
    ]


# ── 2) POST /templates/analyze — 미리보기 (영속 X) ───────────────────────────
@router.post("/analyze", response_model=AnalyzedTemplateResponse)
async def analyze_template(
    request: Request,
    user: Annotated[UserInfo, Depends(get_current_user)],
    file: Annotated[UploadFile, File()],
) -> AnalyzedTemplateResponse:
    """업로드 .xlsx → ADR-006/011 휴리스틱 결과 SheetConfig 미리보기. DB 영속 안 함."""
    _ = user  # IDOR 무관 — 단순 미리보기.
    guard: UploadGuard = request.app.state.upload_guard
    content = await file.read()
    try:
        guard.validate(
            filename=file.filename or "template.xlsx",
            content=content,
            declared_mime=file.content_type
            or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except UploadValidationError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(e),
        ) from e

    try:
        sheets = analyze_workbook(content)
    except TemplateAnalysisError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(e),
        ) from e

    mapping_status = (
        "needs_mapping" if any(not cfg.analyzable for cfg in sheets.values()) else "mapped"
    )
    return AnalyzedTemplateResponse(
        sheets={name: sheet_config_to_view(name, cfg) for name, cfg in sheets.items()},
        mapping_status=mapping_status,
    )


# ── 3) POST /templates — 등록 (분석 + 영속 + 디스크 저장) ────────────────────
@router.post(
    "",
    response_model=TemplateCreatedResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_template(
    request: Request,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
    file: Annotated[UploadFile, File()],
    name: Annotated[str, Form()] = "",
) -> TemplateCreatedResponse:
    """업로드 → 분석 → DB Template + per-user FS 저장."""
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)

    guard: UploadGuard = request.app.state.upload_guard
    content = await file.read()
    try:
        info = guard.validate(
            filename=file.filename or "template.xlsx",
            content=content,
            declared_mime=file.content_type
            or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except UploadValidationError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(e),
        ) from e

    try:
        sheets = analyze_workbook(content)
    except TemplateAnalysisError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(e),
        ) from e

    # SheetConfig dict → JSON 직렬화 (Pydantic model_dump).
    sheets_json: dict[str, Any] = {
        name: sheet_config_to_view(name, cfg).model_dump() for name, cfg in sheets.items()
    }
    mapping_status = (
        "needs_mapping" if any(not cfg.analyzable for cfg in sheets.values()) else "mapped"
    )

    template_name = name or info.original_filename
    # placeholder template_id 로 DB 영속 → 그 후 디스크 저장.
    template = await template_repo.create(
        db,
        user_id=db_user.id,
        name=template_name,
        file_path="",  # 일단 빈 값 — 디스크 저장 후 갱신.
        sheets_json=sheets_json,
        mapping_status=mapping_status,
    )

    file_manager = request.app.state.file_manager
    template_dir = file_manager.template_dir(
        user_oid=user.oid,
        template_id=str(template.id),
        create=True,
    )
    template_path = template_dir / "template.xlsx"
    template_path.write_bytes(content)

    template.file_path = str(template_path)
    await db.commit()

    return TemplateCreatedResponse(
        template_id=template.id,
        name=template.name,
        mapping_status=mapping_status,
    )


# ── 4) GET /templates/{id}/grid — 셀 grid JSON ───────────────────────────────
@router.get("/{template_id}/grid", response_model=GridResponse)
async def get_template_grid(
    template_id: int,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> GridResponse:
    """Templates editor — 셀 값 + 좌표 JSON."""
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    template = await template_repo.get(
        db,
        user_id=db_user.id,
        template_id=template_id,
    )
    p = Path(template.file_path)
    if not p.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="template file missing",
        )
    wb = load_workbook(io.BytesIO(p.read_bytes()), data_only=False)
    sheets: dict[str, GridSheetView] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells: list[GridCell] = []
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                is_formula = isinstance(val, str) and val.startswith("=")
                if isinstance(val, int | float | str):
                    cell_val: str | int | float | None = val
                else:
                    cell_val = str(val)
                cells.append(
                    GridCell(
                        row=row_idx,
                        col=col_idx,
                        value=cell_val,
                        is_formula=is_formula,
                    ),
                )
        sheets[sheet_name] = GridSheetView(
            sheet_name=sheet_name,
            cells=cells,
            max_row=ws.max_row,
            max_col=ws.max_column,
        )
    return GridResponse(sheets=sheets)


# ── 5) PATCH /templates/{id}/cells — 셀 값 일괄 수정 (style deferred) ───────
@router.patch("/{template_id}/cells")
async def patch_template_cells(
    template_id: int,
    body: CellsPatchRequest,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> dict[str, object]:
    """ADR-010 추천 3: 셀 값만 수정. style/병합/줌은 Phase 8+ deferred."""
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    template = await template_repo.get(
        db,
        user_id=db_user.id,
        template_id=template_id,
    )
    p = Path(template.file_path)
    if not p.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="template file missing",
        )

    wb = load_workbook(io.BytesIO(p.read_bytes()))
    updated = 0
    for cell_patch in body.cells:
        if cell_patch.sheet not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=f"sheet '{cell_patch.sheet}' not found",
            )
        ws = wb[cell_patch.sheet]
        ws.cell(row=cell_patch.row, column=cell_patch.col).value = cell_patch.value
        updated += 1

    buf = io.BytesIO()
    wb.save(buf)
    p.write_bytes(buf.getvalue())
    return {"ok": True, "updated_count": updated}


# ── 6) PATCH /templates/{id}/mapping — 매핑 chip override ────────────────────
@router.patch("/{template_id}/mapping")
async def patch_template_mapping(
    template_id: int,
    body: MappingPatchRequest,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> dict[str, object]:
    """매핑 chip 단위 column_map 갱신 — sheets_json 직접 수정."""
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    template = await template_repo.get(
        db,
        user_id=db_user.id,
        template_id=template_id,
    )
    sheets_json: dict[str, Any] = dict(template.sheets_json)
    if body.sheet not in sheets_json:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=f"sheet '{body.sheet}' not in template",
        )
    sheet_view = dict(sheets_json[body.sheet])
    for field_name, value in body.model_dump(exclude_none=True).items():
        if field_name == "sheet":
            continue
        sheet_view[field_name] = value
    sheets_json[body.sheet] = sheet_view
    template.sheets_json = sheets_json
    # 매핑 update → mapping_status='mapped' 자동 (사용자가 명시 매핑 입력함).
    template.mapping_status = "mapped"
    await db.commit()
    return {"ok": True, "sheet": body.sheet}


# ── 7) PATCH /templates/{id} — 메타 (name) ──────────────────────────────────
@router.patch("/{template_id}")
async def patch_template_meta(
    template_id: int,
    body: MetaPatchRequest,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> dict[str, object]:
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    await template_repo.update_meta(
        db,
        user_id=db_user.id,
        template_id=template_id,
        name=body.name,
    )
    await db.commit()
    return {"ok": True}


# ── 8) DELETE /templates/{id} — IDOR 차단 ───────────────────────────────────
@router.delete("/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_template(
    template_id: int,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> None:
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    await template_repo.delete_by_id(db, user_id=db_user.id, template_id=template_id)
    await db.commit()


# ── 9) GET /templates/{id}/raw — 원본 xlsx 다운로드 ──────────────────────────
@router.get("/{template_id}/raw")
async def download_template_raw(
    template_id: int,
    user: Annotated[UserInfo, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(_get_db)],
) -> FileResponse:
    db_user = await user_repo.get_or_create_by_oid(db, oid=user.oid, name=user.name)
    template = await template_repo.get(
        db,
        user_id=db_user.id,
        template_id=template_id,
    )
    p = Path(template.file_path)
    if not p.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="template file missing",
        )
    return FileResponse(path=p, filename=f"{template.name}.xlsx")
