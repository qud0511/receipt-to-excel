import io
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName

from app.schemas.receipt import ReceiptData


def validate_template(xlsx_bytes: bytes) -> list[str]:
    """FIELD_* Named Range 목록을 반환. 없으면 ValueError."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    fields, _ = _field_mapping(wb)
    wb.close()
    if not fields:
        raise ValueError("템플릿에 FIELD_* Named Range가 없습니다.")
    return list(fields.keys())


def analyze_template(xlsx_bytes: bytes) -> dict:
    """Named Range 없는 xlsx에서 시트·헤더 후보·데이터 시작행을 감지."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        candidates: list[dict] = []
        data_start_row: int | None = None

        for row in ws.iter_rows(min_row=1, max_row=25):
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.strip():
                    candidates.append({
                        "row": cell.row,
                        "col": cell.column_letter,
                        "label": val.strip(),
                    })
            if data_start_row is None:
                for cell in row:
                    if isinstance(cell.value, datetime):
                        data_start_row = cell.row
                        break

        if candidates:
            sheets.append({
                "name": sheet_name,
                "candidate_headers": candidates,
                "data_start_row": data_start_row or 6,
            })
    wb.close()
    return {"sheets": sheets}


def inject_named_ranges(
    xlsx_bytes: bytes,
    sheet_name: str,
    field_map: dict[str, str],
    data_start_row: int,
) -> tuple[bytes, list[str]]:
    """field_map 기반으로 FIELD_* / DATA_START Named Range를 xlsx에 주입."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))

    for name in [n for n in wb.defined_names if n.startswith("FIELD_") or n == "DATA_START"]:
        del wb.defined_names[name]

    ws = wb[sheet_name]
    fields: list[str] = []
    safe_sheet = f"'{sheet_name}'" if (" " in sheet_name or "." in sheet_name) else sheet_name

    for field_name, col_letter in field_map.items():
        header_row = 1
        for r in range(1, data_start_row):
            cell = ws[f"{col_letter}{r}"]
            if cell.value is not None:
                header_row = r

        wb.defined_names.add(DefinedName(
            f"FIELD_{field_name}",
            attr_text=f"{safe_sheet}!${col_letter}${header_row}",
        ))
        fields.append(field_name)

    if field_map:
        first_col = next(iter(field_map.values()))
        wb.defined_names.add(DefinedName(
            "DATA_START",
            attr_text=f"{safe_sheet}!${first_col}${data_start_row}",
        ))

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), fields


def build_excel(
    template_path: Path,
    output_path: Path,
    receipts: list[ReceiptData],
) -> None:
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    mapping, target_sheet = _field_mapping(wb)
    start_row = _data_start_row(wb)

    # Named Range가 가리키는 시트에 데이터를 씀
    ws = wb[target_sheet] if target_sheet and target_sheet in wb.sheetnames else wb.active

    # 기존 데이터가 있는 경우 마지막 사용 행 다음부터 쓰기
    write_row = _first_empty_row(ws, start_row, list(mapping.values()))

    for i, receipt in enumerate(receipts):
        row = write_row + i
        row_data = receipt.model_dump()
        for field, col in mapping.items():
            ws.cell(row=row, column=col, value=row_data.get(field))

    wb.save(output_path)
    wb.close()


def _first_empty_row(ws, start_row: int, check_cols: list[int]) -> int:
    """start_row부터 탐색해 매핑 컬럼이 모두 비어 있는 첫 행을 반환."""
    if not check_cols:
        return start_row
    probe = check_cols[:2]  # 첫 두 컬럼만 검사
    row = start_row
    while row <= start_row + 50000:
        if all(ws.cell(row=row, column=c).value is None for c in probe):
            return row
        row += 1
    return start_row  # fallback


def _field_mapping(wb) -> tuple[dict[str, int], str | None]:
    """FIELD_* Named Range에서 {field: col_index} 와 대상 시트명 반환."""
    mapping: dict[str, int] = {}
    target_sheet: str | None = None
    for name in wb.defined_names:
        if name.startswith("FIELD_"):
            field = name[6:]
            sheet_name, cell_ref = list(wb.defined_names[name].destinations)[0]
            col_letter = re.sub(r"[\$\d]", "", cell_ref)
            mapping[field] = column_index_from_string(col_letter)
            target_sheet = sheet_name
    return mapping, target_sheet


def _data_start_row(wb) -> int:
    defined = wb.defined_names
    if "DATA_START" in defined:
        _, cell_ref = list(defined["DATA_START"].destinations)[0]
        return int(re.sub(r"[^\d]", "", cell_ref))

    mapping, _ = _field_mapping(wb)
    if not mapping:
        raise ValueError("템플릿에 FIELD_* Named Range가 없습니다.")

    max_row = 0
    for name in defined:
        if name.startswith("FIELD_"):
            _, cell_ref = list(defined[name].destinations)[0]
            row = int(re.sub(r"[^\d]", "", cell_ref))
            max_row = max(max_row, row)
    return max_row + 1
