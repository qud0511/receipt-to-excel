"""Phase 5.2 — XLSX Writer (R13 동적 행 + formula_cols 보호 + 시트 라우팅).

v1 Bug 1 회귀 차단: ``clear_data_rows`` 가 양식의 기존 더미 데이터 제거.
v1 Bug 2 회귀 차단: ``write_receipt`` 가 ``SheetConfig.formula_cols`` 셀 절대 안 건드림.
R2: 시트 라우팅 — ExpenseRecord.xlsx_sheet (법인/개인) → 해당 시트.
R12: 출력 파일명 `YYYY_MM_지출결의서.xlsx`.
R13: 동적 행 삽입 (Phase 5.2b 별도 — style + merge + formula ref 보존).

CLAUDE.md §"특이사항": SUM 셀은 항상 SUM 수식 — 숫자 직접 쓰기 금지 (formula_cols 보호).
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.template_map import SheetConfig

# regenerate_sum_formulas: `=SUM(F9:F10)` → 새 end_row 로 갱신. column letter 그룹 보존.
_SUM_RANGE_RE = re.compile(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)")

_FALLBACK_CATEGORY = "기타비용"
_OUTPUT_FILENAME_FMT = "{yyyy:04d}_{mm:02d}_지출결의서.xlsx"


def clear_data_rows(ws: Worksheet, config: SheetConfig) -> None:
    """data 영역 (row data_start_row ~ effective_data_end_row) 의 비-formula 셀 클리어.

    v1 Bug 1 회귀 차단. 보호 정책 — 셀 값이 `=` 로 시작하면 skip (수식 셀 보존).
    column-level (formula_cols) 만으로는 부족 — 카테고리 컬럼 (L 등) 은 sum_row 에만 수식이
    있고 data 셀은 작성 가능해야 함. 셀 단위 검사가 정확.
    """
    for row in range(config.data_start_row, config.effective_data_end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col_idx)
            if _is_formula_cell(cell.value):
                continue
            cell.value = None


def write_receipt(
    ws: Worksheet,
    *,
    row_idx: int,
    transaction_date: date,
    merchant: str,
    amount: int,
    expense_column: str,
    sheet_config: SheetConfig,
) -> None:
    """단일 거래 1 행 기입 — 카테고리 매핑 + 기타비용 fallback + 셀 단위 수식 보호.

    카테고리 미매칭 시 기타비용 컬럼 (없으면 첫 category_col) 으로 fallback.
    v1 Bug 2 회귀 차단 — 기존 수식 셀에는 절대 쓰지 않음.
    """
    cfg = sheet_config
    # ── A: 일자 (date 객체 — openpyxl 이 datetime 셀로 저장) ───────────────
    if cfg.date_col:
        _safe_write(ws, cfg.date_col, row_idx, transaction_date)
    # ── B: 가맹점 (AD-1 raw 보존) ─────────────────────────────────────────
    if cfg.merchant_col:
        _safe_write(ws, cfg.merchant_col, row_idx, merchant)
    # ── 카테고리 컬럼 → 금액 ──────────────────────────────────────────────
    target_col = _resolve_category_col(expense_column, cfg)
    if target_col is not None:
        _safe_write(ws, target_col, row_idx, amount)


def _safe_write(ws: Worksheet, col_letter: str, row_idx: int, value: object) -> None:
    """기존 셀 값이 수식이면 skip — v1 Bug 2 셀 단위 보호."""
    cell = ws[f"{col_letter}{row_idx}"]
    if _is_formula_cell(cell.value):
        return
    cell.value = value


def _is_formula_cell(value: object) -> bool:
    """openpyxl 셀의 raw 값이 `=` 로 시작하는 수식 문자열인지."""
    return isinstance(value, str) and value.startswith("=")


def _resolve_category_col(expense_column: str, cfg: SheetConfig) -> str | None:
    """카테고리 → column letter. 미매칭 시 기타비용 → 첫 카테고리 컬럼 fallback."""
    if expense_column in cfg.category_cols:
        return cfg.category_cols[expense_column]
    if _FALLBACK_CATEGORY in cfg.category_cols:
        return cfg.category_cols[_FALLBACK_CATEGORY]
    # 마지막 fallback: 첫 등록된 카테고리 컬럼.
    if cfg.category_cols:
        return next(iter(cfg.category_cols.values()))
    return None


def regenerate_sum_formulas(
    ws: Worksheet,
    config: SheetConfig,
    *,
    new_data_end_row: int,
    sum_row: int | None = None,
) -> None:
    """sum_row 의 ``SUM(F{start}:F{end})`` 패턴을 ``new_data_end_row`` 로 갱신.

    행별 SUM (E열의 ``=SUM(F9:N9)`` 등) 은 단일 행 범위라 영향 없음 — 갱신 대상에서 제외.

    ``sum_row`` 인자: ``insert_row_at`` 후 sum_row 가 shift 되었을 때 명시.
    기본은 ``config.sum_row``.
    """
    target_sum_row = sum_row if sum_row is not None else config.sum_row
    if target_sum_row is None:
        return
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=target_sum_row, column=col_idx)
        val = cell.value
        if not isinstance(val, str):
            continue
        match = _SUM_RANGE_RE.fullmatch(val)
        if not match:
            continue
        start_col, start_row, end_col, _end_row = match.groups()
        # 단일 행 범위 (행별 SUM 같은 경우) skip — `start_row == end_row`.
        if start_row == _end_row:
            continue
        cell.value = f"=SUM({start_col}{start_row}:{end_col}{new_data_end_row})"


def write_workbook(
    template_content: bytes,
    template_sheets: dict[str, SheetConfig],
    rows: list[dict[str, Any]],
    *,
    year: int,
    month: int,
) -> tuple[bytes, str]:
    """전체 잡 orchestration — 시트별 라우팅 → 행별 기입 → 파일명 결정.

    rows[i] 형식: {transaction_date, merchant, amount, expense_column, xlsx_sheet}.
    """
    wb = load_workbook(io.BytesIO(template_content))

    # 시트 kind ("법인"/"개인") → sheet_name 매핑.
    kind_to_sheet: dict[str, str] = {
        cfg.sheet_name: name for name, cfg in template_sheets.items()
    }

    # 시트별 클리어 (rows 가 있든 없든 기존 더미 제거).
    for sheet_name, cfg in template_sheets.items():
        clear_data_rows(wb[sheet_name], cfg)

    # 시트별로 행 카운터.
    sheet_row_idx: dict[str, int] = {
        name: cfg.data_start_row for name, cfg in template_sheets.items()
    }

    for row in rows:
        kind = row["xlsx_sheet"]
        if kind not in kind_to_sheet:
            continue  # 차량 시트 등 범위 외.
        sheet_name = kind_to_sheet[kind]
        cfg = template_sheets[sheet_name]
        write_receipt(
            wb[sheet_name],
            row_idx=sheet_row_idx[sheet_name],
            transaction_date=row["transaction_date"],
            merchant=row["merchant"],
            amount=row["amount"],
            expense_column=row["expense_column"],
            sheet_config=cfg,
        )
        sheet_row_idx[sheet_name] += 1

    # TODO Phase 5.2b: 시트별 sheet_row_idx 가 cfg.effective_data_end_row 초과 시
    # R13 동적 행 삽입 + regenerate_sum_formulas.

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), generate_output_filename(year, month)


def generate_output_filename(year: int, month: int) -> str:
    """R12 파일명 — `YYYY_MM_지출결의서.xlsx` (월 zero-pad)."""
    return _OUTPUT_FILENAME_FMT.format(yyyy=year, mm=month)


# ─────────────────────────────────────────────────────────────────────────────
# R13: 동적 행 삽입 — 스타일 / 병합 / 수식 참조 보존
# ─────────────────────────────────────────────────────────────────────────────


def insert_row_at(
    ws: Worksheet,
    *,
    target_row: int,
    source_row: int,
) -> None:
    """``target_row`` 위치에 1 행 삽입 + ``source_row`` 의 스타일 복제.

    openpyxl ``insert_rows()`` 는 스타일/병합/수식 참조 모두 손실. 본 함수가 수동 보강:
    1. ``ws.insert_rows(target_row)`` 로 row index 이하 모두 +1 shift.
    2. ``source_row`` 의 각 셀 스타일 (font/fill/border/alignment/number_format) 복제.
    3. ``ws.merged_cells.ranges`` 중 target_row 이상의 range 를 수동으로 +1 shift.

    formula 참조 갱신은 별도 ``regenerate_sum_formulas`` 호출. 본 함수는 layout 만 담당.
    """
    from copy import copy

    from openpyxl.worksheet.cell_range import CellRange

    # openpyxl insert_rows 는 셀 값/스타일 일부만 shift. merged_cells 는 수동 처리.
    # 1) 기존 merged ranges 중 target_row 이상의 range 를 +1 shift 위해 미리 수집.
    ranges_to_shift = [
        cr for cr in list(ws.merged_cells.ranges) if cr.min_row >= target_row
    ]

    # 2) shift 대상 merge 해제 → row 삽입 → 새 좌표로 재등록.
    for cr in ranges_to_shift:
        ws.unmerge_cells(str(cr))

    ws.insert_rows(target_row)

    for cr in ranges_to_shift:
        new_range = CellRange(
            min_col=cr.min_col,
            min_row=cr.min_row + 1,
            max_col=cr.max_col,
            max_row=cr.max_row + 1,
        )
        ws.merge_cells(str(new_range))

    # 3) source_row 의 스타일을 target_row 로 복제.
    # ``insert_rows`` 후 source_row 자체는 이동 안 함 (target 이하만 shift).
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        src_cell = ws.cell(row=source_row, column=col_idx)
        dst_cell = ws.cell(row=target_row, column=col_idx)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
