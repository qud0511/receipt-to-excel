"""Phase 6.3 — 카드 사용내역 파서 base + provider 감지.

ADR-010 자료 검증 C-1: UI Upload 의 '법인카드_2025_12.xlsx · 12건' 지원.
영수증 파서 (Phase 4) 와 별도 모듈 — 1 파일 = N 거래.

provider 감지 휴리스틱: 헤더 row 의 한글 keyword set 매칭. 카드사별 다운로드 양식이
다르므로 헤더 시그니처로 판별 (Phase 4 영수증 의 URL ASCII 시그니처 + 텍스트 한글
시그니처 패턴과 동일 사상).
"""

from __future__ import annotations

import io
from typing import Literal

from openpyxl import load_workbook

CardStatementProvider = Literal[
    "shinhan",
    "samsung",
    "hana",
    "woori",
    "hyundai",
    "lotte",
    "kbank",
    "kakaobank",
]


class UnsupportedCardStatementError(ValueError):
    """provider 감지 실패 또는 헤더 미일치 — 422 / 사용자 알림."""


# 헤더 keyword → provider. 카드사별 다운로드 양식이 다르므로 헤더 시그니처로 판별.
# 추후 신규 카드사 추가 = 본 dict 한 줄 + provider 모듈 1 파일 (확장 포인트 3종 패턴).
_HEADER_SIGNATURES: dict[str, tuple[CardStatementProvider, tuple[str, ...]]] = {
    "shinhan_xlsx_v3": ("shinhan", ("거래일자", "거래시각", "가맹점명", "업종", "거래금액")),
    # samsung_xlsx, kbank_xlsx 등 후속 카드사는 본 dict 한 줄 추가 + provider 모듈 1 파일.
}


def detect_provider_from_content(content: bytes, *, suffix: str) -> CardStatementProvider:
    """헤더 row 의 keyword set 으로 provider 결정.

    XLSX: openpyxl 첫 시트 row 1 cell 값.
    CSV: 첫 줄을 split('","') 또는 csv reader.
    """
    headers = _extract_headers(content, suffix=suffix)
    for _name, (provider, sig_keywords) in _HEADER_SIGNATURES.items():
        if all(kw in headers for kw in sig_keywords):
            return provider
    raise UnsupportedCardStatementError(
        f"카드 사용내역 provider 감지 실패 — 알 수 없는 헤더: {headers}"
    )


def _extract_headers(content: bytes, *, suffix: str) -> set[str]:
    """파일 형식별 헤더 row 추출 → str set (공백 트림)."""
    suffix = suffix.lower()
    if suffix == ".xlsx":
        return _extract_xlsx_headers(content)
    if suffix == ".csv":
        return _extract_csv_headers(content)
    raise UnsupportedCardStatementError(f"지원되지 않는 형식: {suffix}")


def _extract_xlsx_headers(content: bytes) -> set[str]:
    """XLSX 첫 시트의 row 1 cell 값 set."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise UnsupportedCardStatementError("XLSX 빈 시트")
    try:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as e:
        raise UnsupportedCardStatementError("XLSX 헤더 row 부재") from e
    return {str(v).strip() for v in first_row if v is not None}


def _extract_csv_headers(content: bytes) -> set[str]:
    """CSV 첫 줄 → str set. UTF-8 only (Phase 6 정책)."""
    import csv as _csv

    text = content.decode("utf-8-sig", errors="strict")
    reader = _csv.reader(io.StringIO(text))
    try:
        first_row = next(reader)
    except StopIteration as e:
        raise UnsupportedCardStatementError("CSV 헤더 row 부재") from e
    return {h.strip() for h in first_row}
