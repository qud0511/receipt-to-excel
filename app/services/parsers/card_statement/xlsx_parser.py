"""XLSX 카드 사용내역 파서 — provider 감지 + row 단위 매핑.

ADR-010 자료 검증 C-1. openpyxl read_only mode 로 대용량 안전 (50MB 한도 +
UploadGuard 검증 후 호출).
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook

from app.domain.parsed_transaction import ParsedTransaction
from app.services.parsers.card_statement.base import (
    UnsupportedCardStatementError,
    detect_provider_from_content,
)
from app.services.parsers.card_statement.providers.shinhan import parse_shinhan_row


def parse_xlsx(content: bytes) -> list[ParsedTransaction]:
    """provider 감지 → 모든 거래 row 매핑.

    빈 시트 또는 알 수 없는 provider 는 ``UnsupportedCardStatementError``.
    """
    provider = detect_provider_from_content(content, suffix=".xlsx")

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        raise UnsupportedCardStatementError("XLSX 빈 시트")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise UnsupportedCardStatementError("XLSX 헤더 row 부재") from e
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    parser = _row_parser_for(provider)
    results: list[ParsedTransaction] = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue  # 빈 row skip.
        row_dict = _zip_headers(headers, row)
        results.append(parser(row_dict))
    return results


def _row_parser_for(
    provider: str,
) -> Any:  # noqa: ANN401 — provider 별 다른 시그니처 대응.
    """provider → row 파서 함수. 신규 카드사 추가 시 본 dispatch 한 줄 추가."""
    if provider == "shinhan":
        return parse_shinhan_row
    raise UnsupportedCardStatementError(f"row parser 미구현: {provider}")


def _zip_headers(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    """헤더 list 와 row tuple zip → dict. openpyxl datetime/date/time → ISO 문자열."""
    result: dict[str, Any] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        val = row[i] if i < len(row) else None
        if isinstance(val, datetime):
            val = val.isoformat()
        elif isinstance(val, date):
            val = val.isoformat()
        elif isinstance(val, time):
            val = val.isoformat()
        result[h] = val
    return result
