import io

from openpyxl import load_workbook

from . import ProcessedInput


def process_spreadsheet(file_bytes: bytes, source_name: str) -> list[ProcessedInput]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = "\t".join(str(v) if v is not None else "" for v in row)
            if cells.strip():
                lines.append(cells)
    wb.close()
    return [ProcessedInput(
        source_name=source_name,
        source_page=0,
        image_b64=None,
        text="\n".join(lines),
        pil_image=None,
    )]
