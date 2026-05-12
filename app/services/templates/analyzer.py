"""Phase 5.1 — 사용자 .xlsx 양식 자동 분석 → SheetConfig 매핑.

ADR-006 §"Phase 5 TemplateAnalyzer 명세" 7 단계 휴리스틱:
1. 시트명 suffix 로 sheet_kind 결정 (`{YY.MM}_법인`, `{YY.MM}_개인`).
2. A2 == "경비 사용 내역서" 시트만 분석 (차량 시트 skip).
3. row 7 keyword 헤더 → category_cols 매핑 (식대/접대비/기타비용/여비교통비 등).
4. Named range (FIELD_*, DATA_START_*) → field-mode 슬롯 채움.
5. A 컬럼 row 9 부터 "합" 포함 셀 스캔 → sum_row.
6. data_end_row = sum_row - 1 (gap 1~2 자동 흡수).
7. formula_cols = data 영역 + sum_row 의 `=` 시작 셀 column letter 집합.

field/category/hybrid 3 modes 는 `SheetConfig.mode` computed_field 가 자동 결정.
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.template_map import SheetConfig

# ADR-011: 시트명 suffix 가정 제거 — UI A사 양식 (Field mode "지출결의서") 등 지원.
# stop word 만 명시 skip (차량 시트 = 별도 양식).
_STOP_SHEET_KEYWORDS = ("차량", "주행", "운행일지")
_ANALYZABLE_A2_MARKERS = ("경비 사용 내역서", "지출결의서", "지출 결의서")
_SUM_LABEL_PATTERN = re.compile(r"^\s*합")  # "합", "합계", "합    계" 모두 매칭.

# ADR-006 의 기존 suffix 휴리스틱은 sheet_kind 결정에만 사용 (분석 대상 판정과 분리).
_SHEET_KIND_SUFFIX_PATTERN = re.compile(r"_(법인|개인)$")

_FIELD_NAMED_RANGE_RE = re.compile(
    r"^(FIELD_(DATE|MERCHANT|TOTAL|PROJECT|NOTE)|DATA_START)_(법인|개인)$"
)

# row 7 / row 8 keyword — Category mode (ADR-006) + Field mode (UI A사 양식, ADR-011).
# 매칭 카운트 ≥ 2 이면 분석 대상 확정 (단일 우연 매치 차단).
_HEADER_KEYWORDS_ROW7 = (
    # ADR-006 Category mode
    "여비교통비",
    "차량유지비",
    "식대",
    "접대비",
    "기타비용",
    "합계",
    "일자",
    "거래처 / 프로젝트명",
    # row 8 sub-headers (Category mode 보강)
    "항공료",
    "버스",
    "숙박비",
    "유류대",
    "주차",
    # ADR-011 Field mode (UI A사 양식)
    "연번",
    "거래일",
    "거래처명",
    "프로젝트명",
    "용도",
    "인원",
    "금액",
    "동석자",
    "비고",
)

# category_cols 추출에 쓰는 keyword (row 7 의 substring 매칭).
_CATEGORY_KEYWORDS = (
    "여비교통비",
    "차량유지비",
    "식대",
    "접대비",
    "기타비용",
    "항공료",
    "버스",
    "숙박비",
    "유류대",
    "주차",
)


class TemplateAnalysisError(ValueError):
    """양식 분석 실패 — 헤더·named range 둘 다 부재 시 raise."""


def analyze_workbook(content: bytes) -> dict[str, SheetConfig]:
    """ADR-011 휴리스틱으로 .xlsx bytes → 시트별 SheetConfig.

    - stop word 시트 (차량/주행/운행일지) → 완전 skip (결과 dict 에 없음).
    - A2 마커 있고 row 7 키워드 카운트 ≥ 2 → 정상 분석 (analyzable=True).
    - A2 마커 있고 row 7 키워드 부족 → analyzable=False (수동 매핑 대기).
    - A2 마커도 없으면 → skip.
    """
    wb = load_workbook(io.BytesIO(content), data_only=False)  # 수식 텍스트 보존.

    result: dict[str, SheetConfig] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if _is_stop_sheet(ws):
            continue
        if not _has_a2_marker(ws):
            continue
        # ADR-011: row 7 키워드 카운트 ≥ 2 OR 시트 scope named range 보유 → 분석 가능.
        # named range 만 가진 field mode 양식 (Phase 5 의 mode="field") 도 흡수.
        sheet_kind_for_named = _extract_sheet_kind(sheet_name)
        has_named_ranges = sheet_kind_for_named is not None and bool(
            _extract_field_slots(wb, sheet_kind_for_named)
        )
        if _row7_keyword_count(ws) >= 2 or has_named_ranges:
            cfg = _analyze_sheet(wb, ws)
            if cfg is not None:
                result[sheet_name] = cfg
        else:
            # A2 마커는 있지만 row 7 헤더 부족 + named range 없음 — 사용자 수동 매핑 대기.
            result[sheet_name] = _make_unanalyzable_placeholder(ws)

    if not result:
        raise TemplateAnalysisError(
            "분석 가능한 시트 부재 — 'A2 마커 (경비 사용 내역서 / 지출결의서)' 필요"
        )
    return result


def _is_stop_sheet(ws: Worksheet) -> bool:
    """차량 / 주행 / 운행일지 시트는 본 phase 범위 밖 — 명시적 skip."""
    title = ws.title or ""
    return any(kw in title for kw in _STOP_SHEET_KEYWORDS)


def _has_a2_marker(ws: Worksheet) -> bool:
    """A2 셀이 분석 가능 마커 중 하나를 포함."""
    a2 = ws["A2"].value
    if not isinstance(a2, str):
        return False
    return any(marker in a2 for marker in _ANALYZABLE_A2_MARKERS)


def _row7_keyword_count(ws: Worksheet) -> int:
    """row 7 의 cell 중 _HEADER_KEYWORDS_ROW7 매칭 수."""
    count = 0
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=7, column=col_idx).value
        if not isinstance(val, str):
            continue
        for kw in _HEADER_KEYWORDS_ROW7:
            if kw in val:
                count += 1
                break
    return count


def _make_unanalyzable_placeholder(ws: Worksheet) -> SheetConfig:
    """A2 마커는 있으나 row 7 헤더 부족 — analyzable=False placeholder.

    sheet_name = 시트 title 그대로 (UI 가 "매핑 필요" flag 노출 시 사용자에게 보임).
    """
    return SheetConfig(
        sheet_name=ws.title or "unknown",
        data_start_row=9,
        data_end_row=9,
        header_row=7,
        analyzable=False,
    )


def _analyze_sheet(wb: Workbook, ws: Worksheet) -> SheetConfig | None:
    """단일 시트 분석 → SheetConfig.

    ADR-011: sheet_name = 시트 title 그대로 (suffix 가정 제거).
    sheet_kind 추출 (법인/개인) 은 named range scope 매칭에만 사용 — 없으면 None.
    """
    sheet_kind = _extract_sheet_kind(ws.title or "")

    # ── named range 기반 field 슬롯 (시트 scope, 법인/개인 양식만 해당) ─────
    field_slots: dict[str, str] = {}
    if sheet_kind is not None:
        field_slots = _extract_field_slots(wb, sheet_kind)

    # ── row 7 + row 8 keyword 헤더 기반 category_cols ─────────────────────
    category_cols = _extract_category_cols(ws)

    # ── ADR-011 Field mode 슬롯 추출 (row 7 header → date_col 등) ──────────
    field_mode_slots = _extract_field_mode_slots_row7(ws)
    # named range 슬롯이 우선 (Category mode), 없으면 row 7 매핑 (Field mode).
    for key, col in field_mode_slots.items():
        field_slots.setdefault(key, col)

    # 헤더·named range 둘 다 부재 → 분석 실패 (analyzable=False placeholder 로 회수).
    if not field_slots and not category_cols:
        return None

    # ── sum_row 탐지 (A 컬럼 "합" 라벨 스캔) ───────────────────────────────
    sum_row = _find_sum_row(ws)
    # data_start_row 는 항상 9 (헤더 row 8 의 다음).
    data_start_row = 9
    data_end_row = (sum_row - 1) if sum_row else ws.max_row

    # ── formula_cols 수집 (행별 SUM + sum_row 세로 SUM) ────────────────────
    formula_cols = _extract_formula_cols(ws, data_start_row, sum_row or data_end_row)

    # ADR-011: sheet_name = 시트 title 그대로, sheet_kind = suffix 추출 (Field mode 면 None).
    return SheetConfig(
        sheet_name=ws.title or "unknown",
        sheet_kind=sheet_kind,
        date_col=field_slots.get("DATE"),
        merchant_col=field_slots.get("MERCHANT"),
        project_col=field_slots.get("PROJECT"),
        total_col=field_slots.get("TOTAL"),
        note_col=field_slots.get("NOTE"),
        category_cols=category_cols,
        formula_cols=formula_cols,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        sum_row=sum_row,
        header_row=7,
        analyzable=True,
    )


def _extract_sheet_kind(title: str) -> str | None:
    """suffix '_법인'/'_개인' 가 있으면 추출, 없으면 None (Field mode 양식)."""
    match = _SHEET_KIND_SUFFIX_PATTERN.search(title)
    return match.group(1) if match else None


# ADR-011 Field mode 슬롯 — row 7 한글 헤더 → SheetConfig 슬롯 매핑.
_ROW7_TO_SLOT: dict[str, str] = {
    "거래일": "DATE",
    "일자": "DATE",
    "거래처명": "MERCHANT",
    "거래처": "MERCHANT",
    "프로젝트명": "PROJECT",
    "프로젝트": "PROJECT",
    "금액": "TOTAL",
    "합계": "TOTAL",
    "비고": "NOTE",
}


def _extract_field_mode_slots_row7(ws: Worksheet) -> dict[str, str]:
    """row 7 header text → Field mode 슬롯 (DATE/MERCHANT/PROJECT/TOTAL/NOTE) → column letter."""
    slots: dict[str, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=7, column=col_idx).value
        if not isinstance(val, str):
            continue
        for keyword, slot in _ROW7_TO_SLOT.items():
            if keyword in val and slot not in slots:
                slots[slot] = get_column_letter(col_idx)
                break
    return slots


def _extract_field_slots(wb: Workbook, sheet_kind: str) -> dict[str, str]:
    """named range FIELD_*_{sheet_kind} → 슬롯 → column letter.

    예: FIELD_DATE_법인 = 'sheet'!$A$9 → {"DATE": "A"}.
    """
    slots: dict[str, str] = {}
    for name, defn in wb.defined_names.items():
        match = _FIELD_NAMED_RANGE_RE.match(name)
        if not match or match.group(3) != sheet_kind:
            continue
        slot = match.group(2) or "DATA_START"
        # defn.value 형식: "'시트명'!$A$9" 또는 "시트명!$A$9"
        addr = defn.value or ""
        col_letter = _extract_col_letter(addr)
        if col_letter is None:
            continue
        # FIELD_DATE → "DATE", DATA_START → 슬롯 등록 안함 (data_start_row 결정만).
        if name.startswith("FIELD_"):
            slots[slot] = col_letter
    return slots


def _extract_col_letter(address: str) -> str | None:
    """`'시트명'!$A$9` 또는 `시트명!$A$9` → "A". 파싱 실패 시 None."""
    # `!` 이후 첫 `$` 그룹의 column letter 추출.
    match = re.search(r"!\$?([A-Z]+)\$?\d+", address)
    return match.group(1) if match else None


def _extract_category_cols(ws: Worksheet) -> dict[str, str]:
    """row 7 + row 8 keyword 헤더 → category_cols (substring 매칭).

    G "버스,택시" vs "버스,택시,대리" 변동 흡수: keyword 가 cell text 의 substring 이면 매칭.
    """
    cols: dict[str, str] = {}
    for row in (7, 8):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row, column=col_idx).value
            if not isinstance(cell_val, str):
                continue
            for kw in _CATEGORY_KEYWORDS:
                if kw in cell_val and kw not in cols:
                    cols[kw] = get_column_letter(col_idx)
                    break
    return cols


def _find_sum_row(ws: Worksheet) -> int | None:
    """A 컬럼 row 9 부터 스캔 — 첫 "합" 포함 셀 row."""
    for row in range(9, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and _SUM_LABEL_PATTERN.match(val):
            return row
    return None


def _extract_formula_cols(ws: Worksheet, data_start_row: int, scan_end_row: int) -> set[str]:
    """data_start_row ~ scan_end_row 의 `=` 시작 셀 column letter 집합."""
    cols: set[str] = set()
    for row in range(data_start_row, scan_end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col_idx).value
            if isinstance(val, str) and val.startswith("="):
                cols.add(get_column_letter(col_idx))
    return cols
