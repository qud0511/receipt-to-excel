"""템플릿 워크북 IO — openpyxl 전담(P8.12: api 에서 openpyxl 직접 import 금지).

라우터는 RawCell/RawSheet 와 TemplateSheetNotFoundError 만 사용 — openpyxl·
pydantic 스키마는 본 모듈 경계를 넘지 않음(CLAUDE.md 코드 구조/도메인-스키마 분리).
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class RawCell:
    row: int
    col: int  # 1-based
    value: str | int | float | None
    is_formula: bool


@dataclass(frozen=True, slots=True)
class RawSheet:
    cells: list[RawCell]
    max_row: int
    max_col: int


class TemplateSheetNotFoundError(ValueError):
    """패치 대상 시트가 워크북에 없음. analyzer.TemplateAnalysisError 와 동일 패턴."""

    def __init__(self, sheet: str) -> None:
        self.sheet = sheet
        super().__init__(f"sheet '{sheet}' not found")


def read_grid(content: bytes) -> dict[str, RawSheet]:
    """워크북 bytes → 시트별 비어있지 않은 셀 grid. 수식은 data_only=False 로 보존."""
    wb = load_workbook(io.BytesIO(content), data_only=False)
    sheets: dict[str, RawSheet] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells: list[RawCell] = []
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                is_formula = isinstance(val, str) and val.startswith("=")
                if isinstance(val, int | float | str):
                    cell_val: str | int | float | None = val
                else:
                    cell_val = str(val)
                cells.append(
                    RawCell(
                        row=row_idx,
                        col=col_idx,
                        value=cell_val,
                        is_formula=is_formula,
                    )
                )
        sheets[sheet_name] = RawSheet(
            cells=cells,
            max_row=ws.max_row,
            max_col=ws.max_column,
        )
    return sheets


def apply_cell_patches(
    content: bytes,
    patches: Sequence[tuple[str, int, int, str | int | float | None]],
) -> tuple[bytes, int]:
    """(sheet,row,col,value) 패치 일괄 적용 → (새 워크북 bytes, 갱신 수).

    시트 부재 시 TemplateSheetNotFoundError. 저장은 전체 검증 통과 후 1회 —
    부분 기록 없음(기존 라우터 동작과 동일: 실패 시 디스크 무변경).
    """
    wb = load_workbook(io.BytesIO(content))
    updated = 0
    for sheet, row, col, value in patches:
        if sheet not in wb.sheetnames:
            raise TemplateSheetNotFoundError(sheet)
        ws = wb[sheet]
        ws.cell(row=row, column=col).value = value
        updated += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), updated
