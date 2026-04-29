"""
지출결의서 샘플 템플릿(Named Range 포함)을 생성합니다.
출력: tests/fixtures/template.xlsx

사용법: python scripts/make_sample_template.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName

Path("tests/fixtures").mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "지출결의서"

ws["A1"] = "지출결의서"
ws["A1"].font = Font(bold=True, size=14)

headers = ["날짜", "업체명", "품목", "금액", "부가세", "결제수단", "비고"]
cols    = ["B",    "C",     "D",   "E",   "F",    "G",     "H"]

for col, header in zip(cols, headers):
    cell = ws[f"{col}3"]
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDEBF7")
    cell.alignment = Alignment(horizontal="center")
    wb.defined_names.add(DefinedName(f"FIELD_{header}", attr_text=f"지출결의서!${col}$3"))

wb.defined_names.add(DefinedName("DATA_START", attr_text="지출결의서!$B$4"))

out = Path("tests/fixtures/template.xlsx")
wb.save(out)
print(f"Template saved: {out}")
print(f"Fields: {headers}")
