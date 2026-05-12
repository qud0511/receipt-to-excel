"""합성 .xlsx 양식 — TemplateAnalyzer / XlsxWriter 단위 테스트용.

ADR-006 §"법인 / 개인 시트 — 공통 layout" 의 휴리스틱 추출 대상을 합성으로 재현.
실 양식은 ``tests/smoke/real_templates/`` (.gitignore) 의 round-trip 통합 테스트에서만 사용.
"""

from __future__ import annotations

import io
from typing import Literal

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


def make_template(
    *,
    yymm: str = "26.05",
    mode: Literal["field", "category", "hybrid"] = "hybrid",
    data_rows: int = 2,
    sum_row_gap: int = 0,
    sheet_kinds: tuple[str, ...] = ("법인", "개인"),
) -> bytes:
    """ADR-006 layout 합성 .xlsx — `mode` 별 헤더/named range 조합.

    - field: named range 만 (DATA_START + FIELD_DATE/FIELD_MERCHANT/FIELD_TOTAL).
    - category: row 7~8 keyword 헤더 (여비교통비/식대/접대비/기타비용).
    - hybrid: 양쪽 동시 — v3 R13 주류 케이스.
    """
    wb = Workbook()
    # 기본 시트 제거 후 sheet_kinds 만 생성.
    default = wb.active
    if default is not None:
        wb.remove(default)

    for kind in sheet_kinds:
        sheet_name = f"{yymm}_{kind}"
        ws = wb.create_sheet(sheet_name)

        # ── 헤더 영역 (row 2/4/6/7/8 — ADR-006 고정 위치) ───────────────────
        ws["A2"] = "경비 사용 내역서"
        ws["A4"] = "주식회사 테라시스"
        ws["A6"] = f"사용카드 : {'법인카드' if kind == '법인' else '개인카드&현금'}"

        # row 7 (1차 헤더) — category mode 시그니처.
        if mode in ("category", "hybrid"):
            ws["A7"] = "일자"
            ws["B7"] = "거래처 / 프로젝트명"
            ws["E7"] = "합계"
            ws["F7"] = "여비교통비"  # F~H 그룹
            ws["I7"] = "차량유지비"  # I~J 그룹
            ws["L7"] = "식대"
            ws["M7"] = "접대비"
            ws["N7"] = "기타비용"
            # row 8 (2차 헤더)
            ws["F8"] = "항공료"
            ws["G8"] = "버스,택시,대리"
            ws["H8"] = "숙박비"
            ws["I8"] = "유류대"
            ws["J8"] = "주차,통행료"

        # ── 데이터 영역 (row 9 ~ 9+data_rows-1) ─────────────────────────────
        data_start = 9
        data_end = data_start + data_rows - 1 if data_rows > 0 else data_start - 1
        for r in range(data_start, data_start + data_rows):
            # 빈 더미 행 — 실제 데이터는 XlsxWriter 가 채움.
            ws.cell(row=r, column=1, value="")
            # E 컬럼 행별 SUM 수식 — formula_cells 검출 대상.
            ws.cell(row=r, column=5, value=f"=SUM(F{r}:N{r})")

        # ── sum_row (data_end + 1 + gap) ────────────────────────────────────
        sum_row = data_end + 1 + sum_row_gap
        ws.cell(row=sum_row, column=1, value="합    계")
        ws.cell(row=sum_row, column=5, value=f"=SUM(F{sum_row}:O{sum_row})")
        for col_letter in ("F", "G", "H", "I", "J", "L", "M", "N"):
            col_idx = ord(col_letter) - ord("A") + 1
            formula = f"=SUM({col_letter}9:{col_letter}{data_end})"
            ws.cell(row=sum_row, column=col_idx, value=formula)

        # ── field mode named ranges ─────────────────────────────────────────
        if mode in ("field", "hybrid"):
            # 절대 참조 + 시트명 quote (openpyxl 명세).
            quoted = f"'{sheet_name}'"
            wb.defined_names[f"FIELD_DATE_{kind}"] = DefinedName(
                f"FIELD_DATE_{kind}", attr_text=f"{quoted}!$A$9"
            )
            wb.defined_names[f"FIELD_MERCHANT_{kind}"] = DefinedName(
                f"FIELD_MERCHANT_{kind}", attr_text=f"{quoted}!$B$9"
            )
            wb.defined_names[f"FIELD_TOTAL_{kind}"] = DefinedName(
                f"FIELD_TOTAL_{kind}", attr_text=f"{quoted}!$E$9"
            )
            wb.defined_names[f"DATA_START_{kind}"] = DefinedName(
                f"DATA_START_{kind}", attr_text=f"{quoted}!$A$9"
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_empty_template() -> bytes:
    """빈 양식 — 헤더 / named range 둘 다 부재 → analyzer 가 ValidationError 던져야."""
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "26.05_법인"
        ws["A2"] = "경비 사용 내역서"
        # row 7 헤더 부재 + named range 부재 → mode 결정 불가.
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
