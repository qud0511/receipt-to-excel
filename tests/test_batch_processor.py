import pytest
from datetime import datetime
from unittest.mock import AsyncMock, MagicMock
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from PIL import Image as PilImage

from app.services.batch_processor import run_job
from app.core.job_manager import InMemoryJobManager
from app.services.ollama_client import ExtractError
from app.services.preprocessor import ProcessedInput
from app.schemas.receipt import ReceiptData
from app.schemas.template import Template


def make_input(name: str, page: int = 0, with_image: bool = False) -> ProcessedInput:
    img = PilImage.new("RGB", (100, 80), color="white") if with_image else None
    return ProcessedInput(
        source_name=name,
        source_page=page,
        docling_text="가맹점명: 테스트\n금액: 1000",
        pil_image=img,
    )


def make_receipt(name: str = "스타벅스") -> ReceiptData:
    return ReceiptData(
        날짜="2024-01-15", 가맹점명=name,
        금액=5500, 부가세=500, 카테고리="기타비용", 결제수단="카드",
    )


def _make_template_xlsx(tmp_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = "날짜"
    ws["C2"] = "가맹점명"
    wb.defined_names.add(DefinedName("FIELD_날짜",   attr_text="Sheet1!$B$2"))
    wb.defined_names.add(DefinedName("FIELD_가맹점명", attr_text="Sheet1!$C$2"))
    (tmp_path / "templates").mkdir(exist_ok=True)
    wb.save(tmp_path / "templates" / "tpl_test.xlsx")


def _make_mock_store(tmp_data_dir) -> MagicMock:
    _make_template_xlsx(tmp_data_dir)
    store = MagicMock()
    store.get = AsyncMock(return_value=Template(
        template_id="tpl_test", name="A",
        fields=["날짜", "가맹점명"], custom_prompt=None,
        created_at=datetime.utcnow(),
    ))
    store.get_config = AsyncMock(return_value=None)
    store.template_path = MagicMock(
        return_value=tmp_data_dir / "templates" / "tpl_test.xlsx"
    )
    return store


async def test_run_job_completes(tmp_data_dir):
    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=2)

    ollama = MagicMock()
    ollama.extract_receipt = AsyncMock(return_value=make_receipt())

    config = MagicMock()
    config.data_dir = tmp_data_dir

    inputs = [make_input("a.jpg", with_image=True), make_input("b.jpg", with_image=True)]
    await run_job("j1", inputs, "tpl_test", mgr, ollama, _make_mock_store(tmp_data_dir), config)

    job = await mgr.get("j1")
    assert job.status == "completed"
    assert job.download_url == "/jobs/j1/result"
    assert job.pdf_url == "/jobs/j1/result/pdf"
    assert job.nup_pdf_url == "/jobs/j1/result/pdf/nup"


async def test_run_job_partial_failure(tmp_data_dir):
    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=2)

    ollama = MagicMock()
    ollama.extract_receipt = AsyncMock(side_effect=Exception("timeout"))

    config = MagicMock()
    config.data_dir = tmp_data_dir

    inputs = [make_input("a.jpg"), make_input("b.jpg")]
    await run_job("j1", inputs, "tpl_test", mgr, ollama, _make_mock_store(tmp_data_dir), config)

    job = await mgr.get("j1")
    assert job.status == "completed"
    assert len(job.failed_files) == 2


async def test_run_job_adds_llm_logs(tmp_data_dir):
    """run_job이 LLM 분석 단계 로그를 기록한다."""
    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=1)

    ollama = MagicMock()
    ollama.extract_receipt = AsyncMock(return_value=make_receipt())

    config = MagicMock()
    config.data_dir = tmp_data_dir

    inputs = [make_input("a.jpg", with_image=True)]
    await run_job("j1", inputs, "tpl_test", mgr, ollama, _make_mock_store(tmp_data_dir), config)

    job = await mgr.get("j1")
    msgs = [e.msg for e in job.logs]
    assert any("LLM" in m for m in msgs)
    assert any("완료" in m for m in msgs)


async def test_run_job_logs_extract_error(tmp_data_dir):
    """ExtractError 시 error 레벨 로그가 기록된다."""
    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=1)

    ollama = MagicMock()
    ollama.extract_receipt = AsyncMock(
        side_effect=ExtractError("날짜", "날짜를 인식할 수 없습니다")
    )

    config = MagicMock()
    config.data_dir = tmp_data_dir

    inputs = [make_input("bad.jpg")]
    await run_job("j1", inputs, "tpl_test", mgr, ollama, _make_mock_store(tmp_data_dir), config)

    job = await mgr.get("j1")
    error_logs = [e for e in job.logs if e.level == "error"]
    assert len(error_logs) >= 1
    assert "bad.jpg" in error_logs[0].msg


async def test_preprocess_and_run_logs_ocr_stage(tmp_data_dir):
    """preprocess_and_run이 OCR 단계 로그를 기록한다."""
    from app.services.batch_processor import preprocess_and_run
    from PIL import Image as PilImage
    import io

    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=1)

    ollama = MagicMock()
    ollama.extract_receipt = AsyncMock(return_value=make_receipt())

    config = MagicMock()
    config.data_dir = tmp_data_dir

    img = PilImage.new("RGB", (100, 80), color="white")
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    file_pairs = [(buf.getvalue(), "test.jpg")]

    await preprocess_and_run(
        "j1", file_pairs, "tpl_test",
        mgr, ollama, _make_mock_store(tmp_data_dir), config
    )

    job = await mgr.get("j1")
    msgs = [e.msg for e in job.logs]
    assert any("OCR" in m for m in msgs)
    assert any("test.jpg" in m for m in msgs)


async def test_run_job_extract_error_goes_to_failed_files(tmp_data_dir):
    """ExtractError 발생 시 failed_files에 기록되고 엑셀 행은 추가되지 않는다."""
    mgr = InMemoryJobManager()
    await mgr.create("j1", template_id="tpl_test", total=2)

    ollama = MagicMock()
    # 첫 번째: ExtractError, 두 번째: 정상
    ollama.extract_receipt = AsyncMock(side_effect=[
        ExtractError("날짜", "날짜를 인식할 수 없습니다"),
        make_receipt("정상업체"),
    ])

    config = MagicMock()
    config.data_dir = tmp_data_dir

    inputs = [make_input("bad.jpg"), make_input("good.jpg")]
    await run_job("j1", inputs, "tpl_test", mgr, ollama, _make_mock_store(tmp_data_dir), config)

    job = await mgr.get("j1")
    assert job.status == "completed"
    assert len(job.failed_files) == 1
    assert "bad.jpg" in job.failed_files[0]
    assert "날짜" in job.failed_files[0]  # ExtractError 메시지 포함
