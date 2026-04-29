import io
from datetime import datetime
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from app.services.excel_mapper import (
    analyze_template, build_excel, inject_named_ranges, validate_template,
)
from app.schemas.receipt import ReceiptData


def make_template_xlsx(with_data_start: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = "날짜"
    ws["C2"] = "업체명"
    ws["D2"] = "금액"
    wb.defined_names.add(DefinedName("FIELD_날짜",   attr_text="Sheet1!$B$2"))
    wb.defined_names.add(DefinedName("FIELD_업체명", attr_text="Sheet1!$C$2"))
    wb.defined_names.add(DefinedName("FIELD_금액",   attr_text="Sheet1!$D$2"))
    if with_data_start:
        ws["A5"] = "data_start_marker"
        wb.defined_names.add(DefinedName("DATA_START", attr_text="Sheet1!$A$5"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_receipt(name: str = "스타벅스", amount: int = 5500) -> ReceiptData:
    return ReceiptData(
        날짜="2024-01-15", 업체명=name, 품목="아메리카노",
        금액=amount, 부가세=500, 결제수단="카드",
    )


def test_validate_template_returns_fields():
    fields = validate_template(make_template_xlsx())
    assert set(fields) == {"날짜", "업체명", "금액"}


def test_validate_template_no_fields_raises():
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="FIELD_"):
        validate_template(buf.getvalue())


def test_build_excel_writes_rows(tmp_path):
    template_path = tmp_path / "template.xlsx"
    template_path.write_bytes(make_template_xlsx())
    output_path = tmp_path / "result.xlsx"

    receipts = [make_receipt("A", 1000), make_receipt("B", 2000)]
    build_excel(template_path, output_path, receipts)

    wb = load_workbook(output_path)
    ws = wb.active
    # DATA_START 없음 → FIELD 최대 행(2) + 1 = 3행부터 데이터
    assert ws.cell(row=3, column=3).value == "A"  # 업체명 C열
    assert ws.cell(row=4, column=3).value == "B"


def test_build_excel_with_data_start(tmp_path):
    template_path = tmp_path / "template.xlsx"
    template_path.write_bytes(make_template_xlsx(with_data_start=True))
    output_path = tmp_path / "result.xlsx"

    build_excel(template_path, output_path, [make_receipt("X", 9999)])

    wb = load_workbook(output_path)
    ws = wb.active
    # DATA_START = 5행
    assert ws.cell(row=5, column=3).value == "X"


def make_plain_xlsx() -> bytes:
    """Named Range 없는 일반 xlsx (실제 지출결의서 스타일)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "경비내역"
    ws["A1"] = "경비 사용 내역서"
    ws["A3"] = "일    자"
    ws["B3"] = "거래처"
    ws["C3"] = "합    계"
    ws["D3"] = "비고"
    ws["A5"] = datetime(2025, 3, 10)  # 첫 데이터 행 — data_start 감지용
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_analyze_template_detects_headers():
    result = analyze_template(make_plain_xlsx())
    assert "sheets" in result
    sheet = result["sheets"][0]
    assert sheet["name"] == "경비내역"
    labels = [c["label"] for c in sheet["candidate_headers"]]
    assert any("일" in l for l in labels)
    assert sheet["data_start_row"] == 5


def test_inject_named_ranges_roundtrip(tmp_path):
    xlsx = make_plain_xlsx()
    injected, fields = inject_named_ranges(
        xlsx,
        sheet_name="경비내역",
        field_map={"날짜": "A", "업체명": "B", "금액": "C"},
        data_start_row=5,
    )
    assert set(fields) == {"날짜", "업체명", "금액"}

    wb = load_workbook(io.BytesIO(injected))
    names = list(wb.defined_names.keys())
    assert "FIELD_날짜" in names
    assert "DATA_START" in names

    # build_excel로 실제 쓰기 가능한지 확인
    tpl = tmp_path / "tpl.xlsx"
    out = tmp_path / "out.xlsx"
    tpl.write_bytes(injected)
    build_excel(tpl, out, [make_receipt("테스트가게", 10000)])
    wb2 = load_workbook(out)
    ws2 = wb2.active
    # A5에 기존 datetime 데이터가 있으므로 _first_empty_row 가 row 6 반환
    assert ws2.cell(row=6, column=2).value == "테스트가게"
