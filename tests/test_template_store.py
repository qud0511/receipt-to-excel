import io
import pytest
from openpyxl import Workbook
from app.services.template_store import TemplateStore


def make_dummy_xlsx() -> bytes:
    from openpyxl.workbook.defined_name import DefinedName
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = "날짜"
    ws["C2"] = "금액"
    wb.defined_names.add(DefinedName("FIELD_날짜", attr_text="Sheet1!$B$2"))
    wb.defined_names.add(DefinedName("FIELD_금액", attr_text="Sheet1!$C$2"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def store(tmp_data_dir):
    return TemplateStore(data_dir=tmp_data_dir)


async def test_init_and_list_empty(store):
    await store.init_db()
    templates = await store.list_all()
    assert templates == []


async def test_save_and_get(store):
    await store.init_db()
    t = await store.save(
        name="지출결의서",
        fields=["날짜", "금액"],
        xlsx_bytes=make_dummy_xlsx(),
    )
    assert t.template_id.startswith("tpl_")
    assert t.name == "지출결의서"
    assert t.has_custom_prompt is False

    fetched = await store.get(t.template_id)
    assert fetched.template_id == t.template_id


async def test_update_prompt(store):
    await store.init_db()
    t = await store.save(name="A", fields=["날짜"], xlsx_bytes=make_dummy_xlsx())
    updated = await store.update_prompt(t.template_id, "새 프롬프트")
    assert updated.custom_prompt == "새 프롬프트"
    assert updated.has_custom_prompt is True


async def test_delete(store):
    await store.init_db()
    t = await store.save(name="A", fields=["날짜"], xlsx_bytes=make_dummy_xlsx())
    await store.delete(t.template_id)
    with pytest.raises(KeyError):
        await store.get(t.template_id)
    assert not store.template_path(t.template_id).exists()


async def test_get_nonexistent_raises(store):
    await store.init_db()
    with pytest.raises(KeyError):
        await store.get("tpl_notexist")
