import io
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


def make_valid_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = "날짜"
    wb.defined_names.add(DefinedName("FIELD_날짜", attr_text="Sheet1!$B$2"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def client(tmp_data_dir, monkeypatch):
    import app.api.deps as deps
    from app.services.template_store import TemplateStore
    store = TemplateStore(data_dir=tmp_data_dir)
    deps._template_store = store

    import asyncio
    asyncio.get_event_loop().run_until_complete(store.init_db())

    from app.main import app
    return TestClient(app)


def test_register_template(client):
    resp = client.post(
        "/templates",
        files={"file": ("t.xlsx", make_valid_template_bytes(), "application/octet-stream")},
        data={"name": "지출결의서"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["template_id"].startswith("tpl_")
    assert "날짜" in data["fields"]


def test_register_invalid_template(client):
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    resp = client.post(
        "/templates",
        files={"file": ("bad.xlsx", buf.getvalue(), "application/octet-stream")},
        data={"name": "bad"},
    )
    assert resp.status_code == 422


def test_list_templates(client):
    client.post(
        "/templates",
        files={"file": ("t.xlsx", make_valid_template_bytes(), "application/octet-stream")},
        data={"name": "A"},
    )
    resp = client.get("/templates")
    assert resp.status_code == 200
    assert len(resp.json()) >= 1


def test_delete_template(client):
    reg = client.post(
        "/templates",
        files={"file": ("t.xlsx", make_valid_template_bytes(), "application/octet-stream")},
        data={"name": "del"},
    )
    tid = reg.json()["template_id"]
    resp = client.delete(f"/templates/{tid}")
    assert resp.status_code == 204
    assert client.get(f"/templates/{tid}").status_code == 404


def test_update_prompt_json(client):
    reg = client.post(
        "/templates",
        files={"file": ("t.xlsx", make_valid_template_bytes(), "application/octet-stream")},
        data={"name": "prompt-test"},
    )
    tid = reg.json()["template_id"]

    # 프롬프트 설정
    resp = client.put(f"/templates/{tid}/prompt",
                      json={"custom_prompt": "커스텀 프롬프트"})
    assert resp.status_code == 200
    assert resp.json()["has_custom_prompt"] is True

    # null로 초기화
    resp = client.put(f"/templates/{tid}/prompt",
                      json={"custom_prompt": None})
    assert resp.status_code == 200
    assert resp.json()["has_custom_prompt"] is False
