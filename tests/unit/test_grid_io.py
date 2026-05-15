"""grid_io 단위 테스트 — openpyxl 워크북 IO (P8.12 리팩터)."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.services.templates.grid_io import (
    RawCell,
    RawSheet,
    TemplateSheetNotFoundError,
    apply_cell_patches,
    read_grid,
)


def _wb_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = 10
    ws["C1"] = "=B1+1"  # 수식
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = 3.5
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_read_grid_extracts_cells_and_detects_formula() -> None:
    sheets = read_grid(_wb_bytes())

    assert set(sheets) == {"Sheet1", "Sheet2"}
    s1 = sheets["Sheet1"]
    assert isinstance(s1, RawSheet)
    by_pos = {(c.row, c.col): c for c in s1.cells}
    assert by_pos[(1, 1)] == RawCell(row=1, col=1, value="name", is_formula=False)
    assert by_pos[(1, 2)] == RawCell(row=1, col=2, value=10, is_formula=False)
    assert by_pos[(1, 3)].value == "=B1+1"
    assert by_pos[(1, 3)].is_formula is True
    assert s1.max_row == 1
    assert s1.max_col == 3
    assert sheets["Sheet2"].cells[0] == RawCell(row=1, col=1, value=3.5, is_formula=False)


def test_read_grid_skips_none_cells() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["C1"] = "y"  # B1 은 None — 건너뜀
    buf = io.BytesIO()
    wb.save(buf)
    sheets = read_grid(buf.getvalue())
    positions = {(c.row, c.col) for c in next(iter(sheets.values())).cells}
    assert (1, 2) not in positions
    assert positions == {(1, 1), (1, 3)}
