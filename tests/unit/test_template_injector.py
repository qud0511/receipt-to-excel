"""Phase 5.1c — Template Injector 3 케이스 (TDD).

`inject_named_ranges` 가 SheetConfig 슬롯을 named range (FIELD_*, DATA_START_*) 로
주입. XLSX writer 가 동일 좌표 다시 계산할 필요 없게 함.

CLAUDE.md §"성능": 양식 등록 시 1회 영속 → 잡 실행 시 재분석 안 함.
"""

from __future__ import annotations

import io

from app.domain.template_map import SheetConfig
from app.services.templates.injector import inject_named_ranges
from openpyxl import load_workbook

from tests.fixtures.synthetic_xlsx import make_template


def _make_sheet_config(
    *, with_date: bool = True, with_merchant: bool = True
) -> SheetConfig:
    return SheetConfig(
        sheet_name="법인",
        date_col="A" if with_date else None,
        merchant_col="B" if with_merchant else None,
        total_col="E",
        category_cols={"식대": "L"},
        formula_cols={"E", "F"},
        data_start_row=9,
        data_end_row=10,
        sum_row=11,
        header_row=7,
    )


def test_inject_named_ranges_creates_field_and_data_start() -> None:
    """주입 후 FIELD_DATE_법인, FIELD_MERCHANT_법인, FIELD_TOTAL_법인, DATA_START_법인 모두 등록."""
    xlsx = make_template(mode="category", sheet_kinds=("법인",))  # named range 부재 상태에서 시작.
    cfg = _make_sheet_config()

    updated = inject_named_ranges(xlsx, sheet_name="26.05_법인", config=cfg)

    wb = load_workbook(io.BytesIO(updated))
    names = set(wb.defined_names.keys())
    assert "FIELD_DATE_법인" in names
    assert "FIELD_MERCHANT_법인" in names
    assert "FIELD_TOTAL_법인" in names
    assert "DATA_START_법인" in names

    # 주입된 named range 의 값 = 'sheet'!$A$9 형식.
    assert "$A$9" in wb.defined_names["FIELD_DATE_법인"].value
    assert "$B$9" in wb.defined_names["FIELD_MERCHANT_법인"].value
    assert "$E$9" in wb.defined_names["FIELD_TOTAL_법인"].value


def test_inject_named_ranges_overwrites_existing() -> None:
    """기존 FIELD_DATE_법인 이 다른 좌표를 가리키면 새 값으로 덮어쓰기."""
    xlsx = make_template(mode="field", sheet_kinds=("법인",))  # 이미 FIELD_DATE_법인=$A$9 존재.
    # 새 config 는 date_col=C (다른 컬럼) 로 설정.
    cfg = SheetConfig(
        sheet_name="법인",
        date_col="C",  # 변경됨.
        merchant_col="B",
        total_col="E",
        category_cols={},
        formula_cols=set(),
        data_start_row=9,
        data_end_row=10,
        sum_row=11,
        header_row=7,
    )

    updated = inject_named_ranges(xlsx, sheet_name="26.05_법인", config=cfg)

    wb = load_workbook(io.BytesIO(updated))
    # 덮어쓰기 결과 — 새 좌표 C9 반영.
    assert "$C$9" in wb.defined_names["FIELD_DATE_법인"].value
    # 그리고 named range 중복 등록 0 (단일 정의).
    field_dates = [n for n in wb.defined_names if n == "FIELD_DATE_법인"]
    assert len(field_dates) == 1


def test_inject_named_ranges_quotes_sheet_names_with_spaces() -> None:
    """시트명에 공백 시 openpyxl 명세 따라 single-quote 필수."""
    # 합성 시트명을 공백 포함으로 만들기 위해 직접 워크북 생성.
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "26.05 법인"  # 공백 포함.
        ws["A2"] = "경비 사용 내역서"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx = buf.getvalue()

    cfg = _make_sheet_config()
    updated = inject_named_ranges(xlsx, sheet_name="26.05 법인", config=cfg)

    wb2 = load_workbook(io.BytesIO(updated))
    addr = wb2.defined_names["FIELD_DATE_법인"].value
    # 공백 포함 시트명은 single-quote 로 감싸짐.
    assert addr.startswith("'26.05 법인'!") or addr.startswith("'26.05 법인'!$"), (
        f"sheet name not quoted: {addr}"
    )
