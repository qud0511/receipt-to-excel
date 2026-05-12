"""Phase 5.2 — XLSX Writer R13 + 시트 라우팅 + 파일명 (v1 Bug 1·2 회귀 차단).

v1 Bug 1: 기존 더미 행이 클리어 안 되어 추출 거래와 함께 혼재.
v1 Bug 2: 카테고리 컬럼 매핑 시 formula_cols (E 등 SUM 열) 덮어쓰기 → 합계 식 깨짐.

R12: 파일명 `YYYY_MM_지출결의서.xlsx`.
R13: 거래 수 > 데이터 영역 시 동적 행 삽입 (Batch 2).
"""

from __future__ import annotations

import io
from datetime import date

from app.services.generators.xlsx_writer import (
    clear_data_rows,
    generate_output_filename,
    regenerate_sum_formulas,
    write_receipt,
    write_workbook,
)
from app.services.templates.analyzer import analyze_workbook
from openpyxl import load_workbook

from tests.fixtures.synthetic_xlsx import make_template


def _load_first_sheet(content: bytes, sheet_name: str = "26.05_법인"):
    """테스트 helper — bytes → (workbook, worksheet)."""
    wb = load_workbook(io.BytesIO(content))
    return wb, wb[sheet_name]


# ── 1) v1 Bug 1 — clear_data_rows 가 기존 더미 제거 ────────────────────────────
def test_clear_data_rows_removes_existing_dummy() -> None:
    xlsx = make_template(mode="hybrid", data_rows=2)
    sheets = analyze_workbook(xlsx)
    cfg = sheets["26.05_법인"]

    wb, ws = _load_first_sheet(xlsx)
    # 더미 데이터 행 채우기 (row 9 = data_start_row).
    ws["A9"] = "dummy-date"
    ws["B9"] = "dummy-merchant"
    ws["L9"] = 99999  # 식대 column

    clear_data_rows(ws, cfg)

    # data 영역 클리어됨 — A9, B9, L9 모두 None.
    assert ws["A9"].value is None
    assert ws["B9"].value is None
    assert ws["L9"].value is None
    # 하지만 formula_cols (E열) 은 보존 — clear_data_rows 가 formula 셀은 건드리지 않음.
    assert ws["E9"].value == "=SUM(F9:N9)"


# ── 2) v1 Bug 2 — write_receipt 는 formula_cols 절대 덮어쓰기 금지 ────────────
def test_write_receipt_does_not_overwrite_formula_col() -> None:
    xlsx = make_template(mode="hybrid")
    sheets = analyze_workbook(xlsx)
    cfg = sheets["26.05_법인"]

    wb, ws = _load_first_sheet(xlsx)
    # E 컬럼은 formula_cols 에 포함 — write_receipt 가 절대 건드리지 않아야.
    original_e9 = ws["E9"].value

    write_receipt(
        ws,
        row_idx=9,
        transaction_date=date(2026, 5, 10),
        merchant="테스트가맹점",
        amount=8900,
        expense_column="식대",
        sheet_config=cfg,
    )

    # E9 의 SUM 수식 보존.
    assert ws["E9"].value == original_e9
    # A/B/L 만 채워짐 — 카테고리 컬럼 매핑 동작.
    assert ws["A9"].value == date(2026, 5, 10)
    assert ws["B9"].value == "테스트가맹점"
    assert ws["L9"].value == 8900


# ── 3) 카테고리 컬럼 매핑 — 식대=L, 접대비=M, 기타비용=N ──────────────────────
def test_write_receipt_routes_to_correct_category_col() -> None:
    xlsx = make_template(mode="hybrid")
    cfg = analyze_workbook(xlsx)["26.05_법인"]
    wb, ws = _load_first_sheet(xlsx)

    write_receipt(
        ws,
        row_idx=9,
        transaction_date=date(2026, 5, 10),
        merchant="A",
        amount=100,
        expense_column="접대비",
        sheet_config=cfg,
    )
    write_receipt(
        ws,
        row_idx=10,
        transaction_date=date(2026, 5, 11),
        merchant="B",
        amount=200,
        expense_column="기타비용",
        sheet_config=cfg,
    )
    assert ws["M9"].value == 100  # 접대비 → M
    assert ws["N10"].value == 200  # 기타비용 → N
    assert ws["L9"].value is None
    assert ws["L10"].value is None


# ── 4) 카테고리 미매칭 → 기타비용 fallback ──────────────────────────────────
def test_write_receipt_falls_back_to_기타비용_col_when_category_missing() -> None:
    xlsx = make_template(mode="hybrid")
    cfg = analyze_workbook(xlsx)["26.05_법인"]
    wb, ws = _load_first_sheet(xlsx)

    # cfg.category_cols 에 없는 카테고리 이름 → 기타비용 (N) 으로 fallback.
    write_receipt(
        ws,
        row_idx=9,
        transaction_date=date(2026, 5, 10),
        merchant="A",
        amount=300,
        expense_column="회의비",  # 없는 카테고리.
        sheet_config=cfg,
    )
    assert ws["N9"].value == 300  # 기타비용 fallback.


# ── 5) regenerate_sum_formulas 가 sum_row 의 SUM(범위) 갱신 ──────────────────
def test_write_receipt_regenerates_sum_formula() -> None:
    """data_end_row 변경 시 sum_row 의 SUM 범위 갱신."""
    xlsx = make_template(mode="hybrid", data_rows=2)
    cfg = analyze_workbook(xlsx)["26.05_법인"]
    wb, ws = _load_first_sheet(xlsx)

    # data_rows=2 → data_end_row=10, sum_row=11. F 컬럼 sum_row 의 원본 수식 = "=SUM(F9:F10)".
    assert ws["F11"].value == "=SUM(F9:F10)"

    # data_end_row 가 15 로 늘었다고 가정 (R13 동적 행 삽입 시뮬레이션).
    regenerate_sum_formulas(ws, cfg, new_data_end_row=15)

    # sum_row 의 SUM 범위가 F9:F15 로 갱신.
    assert ws["F11"].value == "=SUM(F9:F15)"
    assert ws["L11"].value == "=SUM(L9:L15)"


# ── 6/7) 시트 라우팅 — xlsx_sheet="법인" 은 법인 시트, "개인" 은 개인 시트 ────
def test_sheet_routing_법인_card_goes_to_법인_sheet() -> None:
    """write_workbook 통합 — ExpenseRecord 의 xlsx_sheet 가 라우팅 결정."""
    xlsx = make_template(mode="hybrid")
    sheets = analyze_workbook(xlsx)
    rows = [
        {
            "transaction_date": date(2026, 5, 10),
            "merchant": "법인-A",
            "amount": 100,
            "expense_column": "식대",
            "xlsx_sheet": "법인",
        },
        {
            "transaction_date": date(2026, 5, 11),
            "merchant": "개인-B",
            "amount": 200,
            "expense_column": "접대비",
            "xlsx_sheet": "개인",
        },
    ]

    out_bytes, _ = write_workbook(xlsx, sheets, rows, year=2026, month=5)

    out_wb = load_workbook(io.BytesIO(out_bytes))
    # 법인 시트 row 9 = 법인-A.
    assert out_wb["26.05_법인"]["B9"].value == "법인-A"
    # 개인 시트 row 9 = 개인-B.
    assert out_wb["26.05_개인"]["B9"].value == "개인-B"


def test_sheet_routing_개인_card_goes_to_개인_sheet() -> None:
    """xlsx_sheet="개인" 인 행만 개인 시트에 기입 — 다른 시트 빈 상태 유지."""
    xlsx = make_template(mode="hybrid")
    sheets = analyze_workbook(xlsx)
    rows = [
        {
            "transaction_date": date(2026, 5, 11),
            "merchant": "개인-C",
            "amount": 500,
            "expense_column": "식대",
            "xlsx_sheet": "개인",
        }
    ]

    out_bytes, _ = write_workbook(xlsx, sheets, rows, year=2026, month=5)
    out_wb = load_workbook(io.BytesIO(out_bytes))
    # 법인 시트 row 9 비어있음.
    assert out_wb["26.05_법인"]["B9"].value is None
    # 개인 시트 row 9 = 개인-C.
    assert out_wb["26.05_개인"]["B9"].value == "개인-C"


# ── 11) R12 파일명 패턴 ─────────────────────────────────────────────────────
def test_filename_follows_R12_pattern_YYYY_MM_지출결의서_xlsx() -> None:
    """R12: `YYYY_MM_지출결의서.xlsx`."""
    assert generate_output_filename(2026, 5) == "2026_05_지출결의서.xlsx"
    assert generate_output_filename(2026, 12) == "2026_12_지출결의서.xlsx"
    # 월 1자리 시 zero-pad.
    assert generate_output_filename(2025, 3) == "2025_03_지출결의서.xlsx"


# ── 8/9/10) R13 동적 행 삽입 — Phase 5.2b ────────────────────────────────────


def test_dynamic_row_insertion_preserves_style() -> None:
    """data_start_row 의 셀 스타일 (font/fill/border) 이 새 행에 복제."""
    from copy import copy

    from app.services.generators.xlsx_writer import insert_row_at
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    # row 9 에 style 적용.
    ws["A9"].font = Font(name="Arial", size=11, bold=True)
    ws["A9"].fill = PatternFill(start_color="FFFF00", fill_type="solid")
    ws["A9"].border = Border(left=Side(style="thin"))
    # 시트 비교 안정성을 위해 row 10 도 존재해야 — sum_row 가능.
    ws["A10"] = "기존 row 10 (sum_row)"

    insert_row_at(ws, target_row=10, source_row=9)

    # row 10 이 새로 삽입됨 → 기존 row 10 → row 11 로 shift.
    assert ws["A11"].value == "기존 row 10 (sum_row)"
    # row 10 의 스타일이 row 9 (source) 와 동일.
    # openpyxl StyleProxy 비교 — `copy` 후 attribute 비교.
    src_font = copy(ws["A9"].font)
    dst_font = copy(ws["A10"].font)
    assert dst_font.name == src_font.name
    assert dst_font.size == src_font.size
    assert dst_font.bold == src_font.bold


def test_dynamic_row_insertion_preserves_merged_cells() -> None:
    """삽입된 행이 기존 merged_cells 의 row offset 와 정합."""
    from app.services.generators.xlsx_writer import insert_row_at
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # B11:D11 merge (sum_row 의 가맹점 그룹 등 가정).
    ws.merge_cells("B11:D11")
    ws["A11"] = "합계"

    insert_row_at(ws, target_row=10, source_row=9)

    # 기존 row 11 → row 12 로 shift 됨 → merge 도 B12:D12 로 이동.
    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "B12:D12" in merged_ranges
    assert "B11:D11" not in merged_ranges


def test_dynamic_row_insertion_adjusts_formula_references() -> None:
    """행 삽입 후 sum_row 의 SUM(F9:F10) → SUM(F9:F11) 로 확장."""
    xlsx = make_template(mode="hybrid", data_rows=2)
    cfg = analyze_workbook(xlsx)["26.05_법인"]
    wb, ws = _load_first_sheet(xlsx)

    from app.services.generators.xlsx_writer import insert_row_at

    # data_rows=2 → data_end_row=10, sum_row=11. row 10 직전에 1행 삽입.
    insert_row_at(ws, target_row=10, source_row=9)

    # sum_row 는 row 12 로 shift, SUM 범위는 F9:F11 (data_end_row + 1) 로 확장.
    # Note: insert_row_at 은 layout 만 — regenerate_sum_formulas 가 새 sum_row 명시.
    regenerate_sum_formulas(ws, cfg, new_data_end_row=11, sum_row=12)
    # 새 sum_row 는 cfg.sum_row + 1 = 12.
    assert ws["F12"].value == "=SUM(F9:F11)"
